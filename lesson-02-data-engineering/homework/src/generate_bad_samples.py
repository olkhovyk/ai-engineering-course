"""Generate 'bad' enterprise documents — realistic challenges for the ingestion pipeline.

Enterprise reality: documents are messy, broken, mislabeled, and unpredictable.
This script creates samples that simulate real-world problems.
"""

from pathlib import Path

SAMPLES_DIR = Path(__file__).parent.parent / "samples" / "enterprise_challenges"


# ---------------------------------------------------------------------------
# Challenge 1: Corrupted / truncated PDF
# Reality: file transfers fail mid-way, storage corruption, partial downloads
# ---------------------------------------------------------------------------
def create_corrupted_pdf():
    path = SAMPLES_DIR / "corrupted_truncated.pdf"
    # Valid PDF header but truncated mid-stream
    content = b"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Contents 4 0 R
/Resources << /Font << /F1 5 0 R >> >> >>
endobj
4 0 obj
<< /Length 44 >>
stream
BT /F1 12 Tf 50 750 Td (Corrupted doc) Tj ET
endstream
endobj
"""
    # Intentionally truncated — missing xref, trailer, %%EOF
    path.write_bytes(content)
    print(f"  [1] Corrupted PDF (truncated, no xref/trailer): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 2: Image-only PDF (scanned document, no text layer)
# Reality: legal docs, old contracts, faxes saved as PDFs
# ---------------------------------------------------------------------------
def create_image_only_pdf():
    path = SAMPLES_DIR / "scanned_no_text.pdf"
    # Minimal PDF with an image XObject but no text content
    # In real life this would be a rasterized page — here we simulate
    # the key property: extracting text yields empty string
    stream = b"q 100 0 0 100 50 700 cm /Im1 Do Q"
    # 1x1 white pixel JPEG (smallest valid JPEG)
    jpeg = (
        b"\xff\xd8\xff\xe0\x00\x10JFIF\x00\x01\x01\x00\x00\x01\x00\x01\x00\x00"
        b"\xff\xdb\x00C\x00\x08\x06\x06\x07\x06\x05\x08\x07\x07\x07\t\t"
        b"\x08\n\x0c\x14\r\x0c\x0b\x0b\x0c\x19\x12\x13\x0f\x14\x1d\x1a"
        b"\x1f\x1e\x1d\x1a\x1c\x1c $.\' \",#\x1c\x1c(7),01444\x1f\'9=82<.342"
        b"\xff\xc0\x00\x0b\x08\x00\x01\x00\x01\x01\x01\x11\x00"
        b"\xff\xc4\x00\x1f\x00\x00\x01\x05\x01\x01\x01\x01\x01\x01\x00\x00"
        b"\x00\x00\x00\x00\x00\x00\x01\x02\x03\x04\x05\x06\x07\x08\t\n\x0b"
        b"\xff\xc4\x00\xb5\x10\x00\x02\x01\x03\x03\x02\x04\x03\x05\x05\x04"
        b"\x04\x00\x00\x01}\x01\x02\x03\x00\x04\x11\x05\x12!1A\x06\x13Qa\x07"
        b"\x22q\x142\x81\x91\xa1\x08#B\xb1\xc1\x15R\xd1\xf0$3br\x82\t\n\x16"
        b"\x17\x18\x19\x1a%&\'()*456789:CDEFGHIJSTUVWXYZcdefghijstuvwxyz"
        b"\x83\x84\x85\x86\x87\x88\x89\x8a\x92\x93\x94\x95\x96\x97\x98\x99"
        b"\x9a\xa2\xa3\xa4\xa5\xa6\xa7\xa8\xa9\xaa\xb2\xb3\xb4\xb5\xb6\xb7"
        b"\xb8\xb9\xba\xc2\xc3\xc4\xc5\xc6\xc7\xc8\xc9\xca\xd2\xd3\xd4\xd5"
        b"\xd6\xd7\xd8\xd9\xda\xe1\xe2\xe3\xe4\xe5\xe6\xe7\xe8\xe9\xea\xf1"
        b"\xf2\xf3\xf4\xf5\xf6\xf7\xf8\xf9\xfa"
        b"\xff\xda\x00\x08\x01\x01\x00\x00?\x00\xfb\xd2\x8a(\x03\xff\xd9"
    )

    obj1 = b"1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n"
    obj2 = b"2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n"
    obj3 = (
        b"3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
        b"/Contents 4 0 R /Resources << /XObject << /Im1 5 0 R >> >> >>\nendobj\n"
    )
    obj4 = f"4 0 obj\n<< /Length {len(stream)} >>\nstream\n".encode() + stream + b"\nendstream\nendobj\n"
    obj5 = (
        f"5 0 obj\n<< /Type /XObject /Subtype /Image /Width 1 /Height 1 "
        f"/ColorSpace /DeviceGray /BitsPerComponent 8 /Length {len(jpeg)} "
        f"/Filter /DCTDecode >>\nstream\n".encode()
        + jpeg
        + b"\nendstream\nendobj\n"
    )

    objects = [obj1, obj2, obj3, obj4, obj5]
    pdf = b"%PDF-1.4\n"
    offsets = []
    for obj in objects:
        offsets.append(len(pdf))
        pdf += obj

    xref_offset = len(pdf)
    pdf += b"xref\n"
    pdf += f"0 {len(objects) + 1}\n".encode()
    pdf += b"0000000000 65535 f \n"
    for off in offsets:
        pdf += f"{off:010d} 00000 n \n".encode()
    pdf += b"trailer\n"
    pdf += f"<< /Size {len(objects) + 1} /Root 1 0 R >>\n".encode()
    pdf += b"startxref\n"
    pdf += f"{xref_offset}\n".encode()
    pdf += b"%%EOF\n"

    path.write_bytes(pdf)
    print(f"  [2] Image-only PDF (no text layer, needs OCR): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 3: Empty file (0 bytes)
# Reality: failed exports, placeholder files, touch-created files
# ---------------------------------------------------------------------------
def create_empty_files():
    for ext in [".pdf", ".docx", ".html"]:
        path = SAMPLES_DIR / f"empty_file{ext}"
        path.write_bytes(b"")
        print(f"  [3] Empty file (0 bytes): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 4: Wrong extension (PDF content in .html file, HTML in .pdf)
# Reality: users rename files, CMS bugs, email attachment corruption
# ---------------------------------------------------------------------------
def create_wrong_extension_files():
    # HTML content saved as .pdf
    html_as_pdf = SAMPLES_DIR / "actually_html.pdf"
    html_as_pdf.write_text(
        "<html><body><h1>I'm actually HTML!</h1>"
        "<p>Someone saved me with a .pdf extension.</p></body></html>"
    )
    print(f"  [4] Wrong extension (HTML saved as .pdf): {html_as_pdf.name}")

    # PDF header in .html file
    pdf_as_html = SAMPLES_DIR / "actually_pdf.html"
    pdf_as_html.write_bytes(
        b"%PDF-1.4\n% This is actually a PDF file saved with .html extension\n"
    )
    print(f"  [4] Wrong extension (PDF saved as .html): {pdf_as_html.name}")


# ---------------------------------------------------------------------------
# Challenge 5: Mixed / broken encodings
# Reality: legacy systems, Eastern European / CJK documents, BOM issues
# ---------------------------------------------------------------------------
def create_encoding_nightmare():
    # UTF-8 with BOM
    path_bom = SAMPLES_DIR / "utf8_with_bom.html"
    content = (
        '<html><head><meta charset="utf-8"><title>BOM Test</title></head>'
        '<body><p>This file has a UTF-8 BOM marker.</p>'
        '<p>Деякі символи українською мовою.</p></body></html>'
    )
    path_bom.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))
    print(f"  [5] Encoding: UTF-8 with BOM: {path_bom.name}")

    # Windows-1251 (Cyrillic) without proper charset declaration
    path_cp1251 = SAMPLES_DIR / "windows1251_no_charset.html"
    cyrillic_html = (
        "<html><head><title>Звіт</title></head>"
        "<body><p>Цей документ в кодуванні Windows-1251 без мета-тегу charset.</p>"
        "<p>Типова проблема з legacy системами.</p></body></html>"
    )
    path_cp1251.write_bytes(cyrillic_html.encode("cp1251"))
    print(f"  [5] Encoding: Windows-1251 without charset: {path_cp1251.name}")

    # Latin-1 with special chars
    path_latin = SAMPLES_DIR / "latin1_mixed.html"
    latin_html = (
        '<html><head><title>Rapport</title></head>'
        '<body><p>Ger\xe4te\xfcbersicht und Ma\xdfnahmen.</p>'
        '<p>R\xe9sum\xe9 des r\xe9sultats fran\xe7ais.</p></body></html>'
    )
    path_latin.write_bytes(latin_html.encode("latin-1"))
    print(f"  [5] Encoding: Latin-1 with special chars: {path_latin.name}")


# ---------------------------------------------------------------------------
# Challenge 6: Deeply nested / malformed HTML
# Reality: CMS-generated HTML, email HTML, Word-exported HTML
# ---------------------------------------------------------------------------
def create_malformed_html():
    path = SAMPLES_DIR / "malformed_deeply_nested.html"
    # Simulates Word-exported HTML with insane nesting and inline styles
    nested = "<div>" * 50
    nested_close = "</div>" * 50
    html = f"""<html>
<head><title>Enterprise Report Q4</title></head>
<body>
<!-- Exported from Microsoft Word 2016 -->
<div class="WordSection1">
{nested}
<p class="MsoNormal" style="margin-bottom:0in;margin-bottom:.0001pt;
line-height:normal;mso-layout-grid-align:none;text-autospace:none">
<span style="font-size:10.0pt;font-family:&quot;Courier New&quot;;
mso-fareast-font-family:&quot;Times New Roman&quot;;color:black">
Important quarterly data buried under 50 levels of nesting.
Revenue increased by 15% year-over-year.
</span></p>
{nested_close}
</div>

<!-- Unclosed tags -->
<table>
<tr><td>Metric<td>Value
<tr><td>Revenue<td>$1.2M
<tr><td>Users<td>45,000
</table>

<!-- Broken entities -->
<p>Profit &amp; Loss for Q4 &mdash; results are &copy; confidential</p>
<p>Temperature was 72&deg;F &amp the margin was &lt;5%</p>

<!-- Mixed content -->
<p>Contact: <a href="mailto:reports@company.com">reports@company.com</p></a>
</body></html>"""
    path.write_text(html)
    print(f"  [6] Malformed HTML (deep nesting, unclosed tags, broken entities): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 7: HTML with excessive boilerplate (low signal-to-noise)
# Reality: scraped web pages, email newsletters, SharePoint exports
# ---------------------------------------------------------------------------
def create_boilerplate_html():
    path = SAMPLES_DIR / "boilerplate_heavy.html"
    nav_items = "\n".join(
        f'        <li><a href="/page{i}">Navigation Item {i}</a></li>'
        for i in range(30)
    )
    sidebar_items = "\n".join(
        f'        <div class="widget"><h4>Widget {i}</h4><p>Sidebar content {i}</p></div>'
        for i in range(20)
    )
    html = f"""<!DOCTYPE html>
<html>
<head>
    <title>Company Intranet - Important Policy Update</title>
    <style>
        /* 200 lines of CSS that someone copy-pasted from Bootstrap */
        .container {{ max-width: 1200px; }} .row {{ display: flex; }}
        .col {{ flex: 1; }} .nav {{ list-style: none; }}
        /* ... imagine 200 more lines ... */
    </style>
    <script>
        // Analytics, tracking, cookie consent — all noise
        var _gaq = _gaq || [];
        _gaq.push(['_setAccount', 'UA-XXXXX-X']);
        function initCookieBanner() {{ /* ... */ }}
        function trackEvent(cat, action) {{ /* ... */ }}
    </script>
</head>
<body>
    <header>
        <nav>
            <ul>
{nav_items}
            </ul>
        </nav>
    </header>

    <div class="container">
        <div class="row">
            <main class="col-8">
                <!-- THE ACTUAL CONTENT (only 3 paragraphs in a sea of boilerplate) -->
                <h1>Important Policy Update: Data Handling Procedures</h1>
                <p>Effective March 1, 2026, all departments must follow updated data handling
                procedures for personally identifiable information (PII). This includes new
                requirements for encryption at rest and in transit.</p>
                <p>Key changes: 1) All PII must be encrypted using AES-256. 2) Data retention
                period reduced from 7 years to 3 years. 3) Access logs must be reviewed monthly.</p>
                <p>Non-compliance may result in disciplinary action. Contact the Data Protection
                Officer at dpo@company.com for questions.</p>
            </main>
            <aside class="col-4">
{sidebar_items}
            </aside>
        </div>
    </div>

    <footer>
        <p>&copy; 2026 Company Inc. All rights reserved.</p>
        <p>Legal | Privacy | Terms | Cookie Settings | Accessibility | Sitemap</p>
        <p>123 Corporate Blvd, Suite 400, Business City, ST 12345</p>
    </footer>

    <script src="analytics.js"></script>
    <script src="chat-widget.js"></script>
    <script src="cookie-consent.js"></script>
</body>
</html>"""
    path.write_text(html)
    print(f"  [7] Boilerplate-heavy HTML (low signal-to-noise): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 8: Password-protected / encrypted PDF
# Reality: HR docs, financial reports, legal contracts
# ---------------------------------------------------------------------------
def create_password_protected_pdf():
    path = SAMPLES_DIR / "password_protected.pdf"
    # Simulates an encrypted PDF — parsers should detect and report this
    # Real encrypted PDFs have /Encrypt dict; we create a minimal one
    content = b"""%PDF-1.6
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] >>
endobj
4 0 obj
<< /Filter /Standard /V 4 /R 4 /Length 128
/O <abcdef1234567890abcdef1234567890abcdef1234567890abcdef1234567890>
/U <abcdef1234567890abcdef1234567890abcdef1234567890abcdef1234567890>
/P -3904 >>
endobj
"""
    pdf = content
    xref_offset = len(pdf)
    pdf += b"xref\n0 5\n"
    pdf += b"0000000000 65535 f \n"
    pdf += b"0000000009 00000 n \n"
    pdf += b"0000000058 00000 n \n"
    pdf += b"0000000115 00000 n \n"
    pdf += b"0000000198 00000 n \n"
    pdf += b"trailer\n"
    pdf += b"<< /Size 5 /Root 1 0 R /Encrypt 4 0 R >>\n"
    pdf += b"startxref\n"
    pdf += f"{xref_offset}\n".encode()
    pdf += b"%%EOF\n"

    path.write_bytes(pdf)
    print(f"  [8] Password-protected PDF (encrypted): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 9: Huge repetitive document (memory pressure)
# Reality: auto-generated logs, database exports, audit trails
# ---------------------------------------------------------------------------
def create_huge_html():
    path = SAMPLES_DIR / "huge_repetitive.html"
    rows = "\n".join(
        f"<tr><td>{i}</td><td>Transaction {i}</td><td>${i * 3.14:.2f}</td>"
        f"<td>2026-01-{(i % 28) + 1:02d}</td><td>{'Active' if i % 3 else 'Closed'}</td></tr>"
        for i in range(1, 10001)
    )
    html = f"""<html><head><title>Audit Log Export</title></head><body>
<h1>Transaction Audit Log — 10,000 Records</h1>
<table>
<tr><th>ID</th><th>Description</th><th>Amount</th><th>Date</th><th>Status</th></tr>
{rows}
</table></body></html>"""
    path.write_text(html)
    size_mb = path.stat().st_size / 1024 / 1024
    print(f"  [9] Huge HTML ({size_mb:.1f} MB, 10K rows): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 10: Mixed-language document
# Reality: multinational orgs, UN documents, EU regulations
# ---------------------------------------------------------------------------
def create_multilingual_html():
    path = SAMPLES_DIR / "multilingual.html"
    html = """<!DOCTYPE html>
<html lang="mul">
<head><meta charset="utf-8"><title>Multilingual Report</title></head>
<body>
<h1>Quarterly Report / Щоквартальний звіт / 四半期報告</h1>

<h2>English</h2>
<p>Revenue grew by 12% in Q4 2025, driven primarily by expansion into
new markets in Southeast Asia and Eastern Europe.</p>

<h2>Українська</h2>
<p>Дохід зріс на 12% у четвертому кварталі 2025 року, переважно завдяки
розширенню на нові ринки Південно-Східної Азії та Східної Європи.</p>

<h2>日本語</h2>
<p>2025年第4四半期の収益は12%増加し、主に東南アジアと東ヨーロッパの
新市場への拡大が牽引しました。</p>

<h2>العربية</h2>
<p dir="rtl">نمت الإيرادات بنسبة 12٪ في الربع الرابع من عام 2025، مدفوعة
بشكل أساسي بالتوسع في أسواق جديدة في جنوب شرق آسيا وأوروبا الشرقية.</p>

<h2>Deutsch</h2>
<p>Der Umsatz stieg im vierten Quartal 2025 um 12%, hauptsächlich getrieben
durch die Expansion in neue Märkte in Südostasien und Osteuropa.</p>
</body></html>"""
    path.write_text(html, encoding="utf-8")
    print(f"  [10] Multilingual document (5 languages inc. RTL): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 11: DOCX-like broken ZIP (DOCX is just a ZIP)
# Reality: email attachment corruption, partial sync, disk errors
# ---------------------------------------------------------------------------
def create_broken_docx():
    path = SAMPLES_DIR / "broken_archive.docx"
    # DOCX is a ZIP file. This has a valid ZIP header but corrupted contents
    # PK\x03\x04 is the ZIP local file header signature
    path.write_bytes(
        b"PK\x03\x04\x14\x00\x00\x00\x08\x00"
        b"\x00\x00\x00\x00\x00\x00\x00\x00"
        b"\x00\x00\x00\x00\x00\x00\x00\x00"
        b"word/document.xml"  # file name
        b"\x00" * 100  # corrupted compressed data
    )
    print(f"  [11] Broken DOCX (corrupted ZIP archive): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 12: Binary file with document extension
# Reality: database dumps, images renamed, system files
# ---------------------------------------------------------------------------
def create_binary_as_document():
    path = SAMPLES_DIR / "binary_garbage.pdf"
    import os
    path.write_bytes(os.urandom(1024))
    print(f"  [12] Random binary data with .pdf extension: {path.name}")


# ---------------------------------------------------------------------------
# Challenge 13: Normal Excel (for baseline comparison)
# Reality: most enterprise data lives in Excel
# ---------------------------------------------------------------------------
def create_normal_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        print("  [13] Skipped (install openpyxl)")
        return

    path = SAMPLES_DIR / "normal_report.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Q4 Revenue"
    ws.append(["Region", "Q1", "Q2", "Q3", "Q4", "Total"])
    data = [
        ("North America", 1200000, 1350000, 1100000, 1500000),
        ("Europe", 800000, 920000, 870000, 1050000),
        ("APAC", 450000, 520000, 610000, 780000),
        ("LATAM", 200000, 230000, 250000, 310000),
    ]
    for region, q1, q2, q3, q4 in data:
        ws.append([region, q1, q2, q3, q4, q1 + q2 + q3 + q4])

    ws2 = wb.create_sheet("Notes")
    ws2.append(["Date", "Note"])
    ws2.append(["2025-12-15", "APAC expansion completed ahead of schedule"])
    ws2.append(["2026-01-10", "LATAM currency impact adjusted"])

    wb.save(str(path))
    print(f"  [13] Normal Excel (baseline, 2 sheets): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 14: Excel with merged cells, formulas, multiple sheets
# Reality: finance team reports with complex formatting
# ---------------------------------------------------------------------------
def create_complex_xlsx():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("  [14] Skipped (install openpyxl)")
        return

    path = SAMPLES_DIR / "complex_merged_formulas.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget FY2026"

    # Merged header cells (common in enterprise reports)
    ws.merge_cells("A1:F1")
    ws["A1"] = "Annual Budget Report — CONFIDENTIAL"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = "Fiscal Year 2026 — All amounts in USD"

    # Headers
    ws.append([])  # row 3 blank
    ws.append(["Department", "Q1 Budget", "Q1 Actual", "Q2 Budget", "Q2 Actual", "Variance"])
    ws[4][0].font = Font(bold=True)

    departments = [
        ("Engineering", 500000, 480000, 520000, 535000),
        ("Marketing", 300000, 310000, 280000, 295000),
        ("Sales", 400000, 420000, 450000, 430000),
        ("Operations", 200000, 195000, 210000, 225000),
        ("HR", 150000, 145000, 160000, 155000),
    ]
    for i, (dept, q1b, q1a, q2b, q2a) in enumerate(departments, start=5):
        ws.append([dept, q1b, q1a, q2b, q2a])
        # Formula for variance
        ws.cell(row=i, column=6, value=f"=(C{i}-B{i})+(E{i}-D{i})")

    # Totals row with SUM formulas
    total_row = 5 + len(departments)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col in range(2, 7):
        col_letter = chr(64 + col)
        ws.cell(row=total_row, column=col,
                value=f"=SUM({col_letter}5:{col_letter}{total_row - 1})")

    # Sheet 2: hidden sheet (common in enterprise — raw data sheets)
    ws2 = wb.create_sheet("Raw Data (DO NOT EDIT)")
    for i in range(1, 101):
        ws2.append([f"TXN-{i:04d}", f"Vendor {i % 10}", i * 123.45,
                     "2026-01-15", "approved" if i % 5 else "pending"])

    # Sheet 3: mostly empty (sparse data)
    ws3 = wb.create_sheet("Sparse Notes")
    ws3["A1"] = "Notes"
    ws3["A50"] = "Important note buried in row 50"
    ws3["C100"] = "Another note far away at C100"
    ws3["Z1"] = "Column Z header"

    wb.save(str(path))
    print(f"  [14] Complex Excel (merged cells, formulas, 3 sheets): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 15: Corrupted Excel (broken ZIP, like broken DOCX)
# Reality: email corruption, partial downloads
# ---------------------------------------------------------------------------
def create_corrupted_xlsx():
    path = SAMPLES_DIR / "corrupted.xlsx"
    # XLSX is ZIP — same trick as broken DOCX
    path.write_bytes(
        b"PK\x03\x04\x14\x00\x00\x00\x08\x00"
        b"\x00\x00\x00\x00\x00\x00\x00\x00"
        b"\x00\x00\x00\x00\x00\x00\x00\x00"
        b"xl/worksheets/sheet1.xml"
        b"\x00" * 100
    )
    print(f"  [15] Corrupted Excel (broken ZIP): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 16: Huge Excel (10K+ rows, memory pressure)
# Reality: database exports, log dumps, CRM extracts
# ---------------------------------------------------------------------------
def create_huge_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [16] Skipped (install openpyxl)")
        return

    path = SAMPLES_DIR / "huge_export.xlsx"
    wb = Workbook(write_only=True)  # write_only for memory efficiency
    ws = wb.create_sheet("Export")
    ws.append(["ID", "Customer", "Email", "Amount", "Date", "Status", "Region"])

    regions = ["NA", "EU", "APAC", "LATAM", "MEA"]
    statuses = ["active", "churned", "trial", "suspended"]
    for i in range(1, 10001):
        ws.append([
            i,
            f"Customer {i}",
            f"user{i}@company{i % 50}.com",
            round(i * 3.14 + (i % 100) * 7.5, 2),
            f"2026-{(i % 12) + 1:02d}-{(i % 28) + 1:02d}",
            statuses[i % len(statuses)],
            regions[i % len(regions)],
        ])

    wb.save(str(path))
    size_kb = path.stat().st_size / 1024
    print(f"  [16] Huge Excel ({size_kb:.0f} KB, 10K rows): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 17: Empty Excel (has structure but no data)
# Reality: template files, just-created reports
# ---------------------------------------------------------------------------
def create_empty_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [17] Skipped (install openpyxl)")
        return

    path = SAMPLES_DIR / "empty_with_headers.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append(["ID", "Name", "Value", "Date", "Status"])
    # No data rows — only headers
    wb.save(str(path))
    print(f"  [17] Empty Excel (headers only, no data): {path.name}")


# ---------------------------------------------------------------------------
# Challenge 18: Excel with special characters and Unicode
# Reality: international customer data, CJK names, Arabic text
# ---------------------------------------------------------------------------
def create_unicode_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        print("  [18] Skipped (install openpyxl)")
        return

    path = SAMPLES_DIR / "unicode_data.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "International Clients"
    ws.append(["Client", "Country", "Revenue", "Notes"])
    ws.append(["Müller GmbH", "Germany", 450000, "Geräteübersicht Q4"])
    ws.append(["Тест Компанія", "Ukraine", 120000, "Щоквартальний звіт"])
    ws.append(["田中太郎株式会社", "Japan", 890000, "四半期報告書"])
    ws.append(["شركة الاختبار", "UAE", 670000, "التقرير الربعي"])
    ws.append(["Société Générale", "France", 1200000, "Résumé financier"])
    ws.append(["Empresa São Paulo", "Brazil", 340000, "Relatório trimestral"])
    ws.append(["한국테스트", "South Korea", 560000, "분기 보고서"])

    wb.save(str(path))
    print(f"  [18] Unicode Excel (7 languages/scripts): {path.name}")


def main():
    SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
    print("Generating enterprise challenge documents...\n")

    create_corrupted_pdf()
    create_image_only_pdf()
    create_empty_files()
    create_wrong_extension_files()
    create_encoding_nightmare()
    create_malformed_html()
    create_boilerplate_html()
    create_password_protected_pdf()
    create_huge_html()
    create_multilingual_html()
    create_broken_docx()
    create_binary_as_document()
    create_normal_xlsx()
    create_complex_xlsx()
    create_corrupted_xlsx()
    create_huge_xlsx()
    create_empty_xlsx()
    create_unicode_xlsx()

    print(f"\nDone! Created {len(list(SAMPLES_DIR.iterdir()))} challenge files in {SAMPLES_DIR}")
    print("\nRun the pipeline on them:")
    print(f"  python src/main.py --input {SAMPLES_DIR}")


if __name__ == "__main__":
    main()
